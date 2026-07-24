import os
import sys
from copy import copy

import openpyxl
from openpyxl.worksheet.cell_range import CellRange


def _template_path():
    if getattr(sys, 'frozen', False):
        base = getattr(sys, '_MEIPASS', os.path.dirname(sys.executable))
        return os.path.join(base, 'templates', 'quote_template.xlsx')
    return os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        'templates', 'quote_template.xlsx')


TEMPLATE = None
FIRST_ITEM_ROW = 4
TEMPLATE_ITEM_ROWS = 8
TOTAL_ROW = 12


def _shift_merges_and_heights(ws, at, count):
    ranges = [str(r) for r in ws.merged_cells.ranges]
    below = []
    for r in ranges:
        cr = CellRange(r)
        if cr.min_row >= at:
            below.append(r)
            ws.unmerge_cells(r)
    max_row = ws.max_row
    heights = {r: ws.row_dimensions[r].height
               for r in range(at, max_row + 1)
               if r in ws.row_dimensions and ws.row_dimensions[r].height}
    ws.insert_rows(at, count)
    for r, h in heights.items():
        ws.row_dimensions[r + count].height = h
    for r in below:
        cr = CellRange(r)
        cr.min_row += count
        cr.max_row += count
        ws.merge_cells(str(cr))


def _copy_row_style(ws, src_row, dst_row, max_col=9):
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        dst._style = copy(src._style)
    if src_row in ws.row_dimensions and ws.row_dimensions[src_row].height:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def export_quote(quote, items, save_path):
    wb = openpyxl.load_workbook(_template_path())
    ws = wb.active

    n = len(items)
    extra = max(0, n - TEMPLATE_ITEM_ROWS)
    if extra:
        _shift_merges_and_heights(ws, TOTAL_ROW, extra)
        for i in range(extra):
            _copy_row_style(ws, TOTAL_ROW - 1, TOTAL_ROW + i)

    d = quote['quote_date']
    try:
        y, m, day = d.split('-')
        date_cn = '{}年{}月{}日'.format(int(y), int(m), int(day))
    except Exception:
        date_cn = d
    ws['A2'] = ('报价日期：{}（报价有效期30天）\u3000\u3000\u3000项目名称：{}\n'
                .format(date_cn, quote['project_name']) + '\u3000' * 27)

    for i in range(TEMPLATE_ITEM_ROWS + extra):
        r = FIRST_ITEM_ROW + i
        if i < n:
            it = items[i]
            ws.cell(r, 1, i + 1)
            ws.cell(r, 2, it.get('name', ''))
            ws.cell(r, 3, it.get('spec', ''))
            ws.cell(r, 4, it.get('code', ''))
            ws.cell(r, 5, it.get('price', 0))
            ws.cell(r, 6, it.get('unit', ''))
            ws.cell(r, 7, it.get('qty', 0))
            ws.cell(r, 8, '=E{}*G{}'.format(r, r))
            ws.cell(r, 9, it.get('remark', ''))
        else:
            for col in range(1, 10):
                ws.cell(r, col).value = None

    total_row = TOTAL_ROW + extra
    last_item_row = FIRST_ITEM_ROW + TEMPLATE_ITEM_ROWS + extra - 1
    ws.cell(total_row, 3).value = '=H{}'.format(total_row)
    ws.cell(total_row, 8).value = '=SUM(H{}:H{})'.format(FIRST_ITEM_ROW, last_item_row)

    ws.cell(total_row + 1, 1).value = quote.get('plan', '')
    ws.cell(total_row + 2, 1).value = quote.get('seller', '')
    ws.cell(total_row + 2, 5).value = quote.get('buyer', '')

    wb.save(save_path)
    return save_path
